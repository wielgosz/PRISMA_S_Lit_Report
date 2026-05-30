"""Input workbook loader and validator for the v2.2 Desktop Runner.

The desktop runner treats the Excel input workbook as the editable source of
truth. This module reads the user-editable tabs, validates required fields,
exports runtime CSV copies, and creates a runtime copy of the v1.3 baseline
workbook where A1/B1 metadata sheets have been replaced by the user's input
workbook. The baseline workbook's Text_Extraction_QA and Document_Term_Matrix
sheets remain intact for reproducibility checks.
"""
from __future__ import annotations

import csv
import json
import shutil
from dataclasses import dataclass, asdict
from datetime import datetime
from pathlib import Path
from typing import Dict, Iterable, List, Optional, Tuple

from openpyxl import load_workbook
from openpyxl.workbook.workbook import Workbook

REQUIRED_SHEETS = [
    "A1_Organizations",
    "B1_Corpus_Documents",
    "Dictionary",
    "Run_Settings",
    "Exclusions_Duplicates",
]

REQUIRED_A1_COLUMNS = ["org_id", "canonical_org_name"]
REQUIRED_B1_COLUMNS = ["doc_id", "file_name", "title", "publishing_org_id", "apa_reference"]
REQUIRED_DICTIONARY_COLUMNS = [
    "category", "term_id", "canonical_term", "search_variant",
    "variant_type", "roll_up_to_canonical", "active",
]

@dataclass
class ValidationIssue:
    severity: str
    sheet: str
    row: str
    field: str
    message: str


def _sheet_to_rows(wb: Workbook, sheet_name: str) -> List[Dict[str, str]]:
    ws = wb[sheet_name]
    # Find the first non-empty row and treat it as the header.
    header_row = None
    headers: List[str] = []
    for idx, row in enumerate(ws.iter_rows(values_only=True), start=1):
        vals = ["" if v is None else str(v).strip() for v in row]
        if any(vals):
            # Skip one-cell intro rows.
            if len([v for v in vals if v]) <= 1:
                continue
            header_row = idx
            headers = vals
            break
    if not header_row:
        return []
    rows: List[Dict[str, str]] = []
    for row_idx, row in enumerate(ws.iter_rows(min_row=header_row+1, values_only=True), start=header_row+1):
        vals = ["" if v is None else str(v).strip() for v in row]
        if not any(vals):
            continue
        rec = {headers[i]: vals[i] if i < len(vals) else "" for i in range(len(headers)) if headers[i]}
        rec["__excel_row"] = str(row_idx)
        rows.append(rec)
    return rows


def read_input_workbook(path: Path) -> Dict[str, List[Dict[str, str]]]:
    wb = load_workbook(path, data_only=False, read_only=True)
    missing = [s for s in REQUIRED_SHEETS if s not in wb.sheetnames]
    if missing:
        raise ValueError(f"Input workbook is missing required sheet(s): {', '.join(missing)}")
    data = {sheet: _sheet_to_rows(wb, sheet) for sheet in REQUIRED_SHEETS if sheet in wb.sheetnames}
    # Optional tabs.
    for sheet in ["New_Documents", "Validation_Log"]:
        if sheet in wb.sheetnames:
            data[sheet] = _sheet_to_rows(wb, sheet)
    wb.close()
    return data


def _columns_present(rows: List[Dict[str, str]], required: Iterable[str]) -> Tuple[bool, List[str]]:
    if not rows:
        return False, list(required)
    cols = set(rows[0].keys())
    missing = [c for c in required if c not in cols]
    return len(missing) == 0, missing


def validate_input_data(data: Dict[str, List[Dict[str, str]]], pdf_folder: Optional[Path] = None) -> List[ValidationIssue]:
    issues: List[ValidationIssue] = []

    checks = [
        ("A1_Organizations", REQUIRED_A1_COLUMNS),
        ("B1_Corpus_Documents", REQUIRED_B1_COLUMNS),
        ("Dictionary", REQUIRED_DICTIONARY_COLUMNS),
    ]
    for sheet, required in checks:
        ok, missing = _columns_present(data.get(sheet, []), required)
        if not ok:
            issues.append(ValidationIssue("ERROR", sheet, "header", ", ".join(missing), f"Missing required columns: {', '.join(missing)}"))

    a1 = data.get("A1_Organizations", [])
    b1 = data.get("B1_Corpus_Documents", [])
    dictionary = data.get("Dictionary", [])
    exclusions = data.get("Exclusions_Duplicates", [])

    org_ids = {r.get("org_id", "").strip() for r in a1 if r.get("org_id", "").strip()}
    for r in b1:
        row = r.get("__excel_row", "")
        if r.get("file_name", "").strip() and not r.get("publishing_org_id", "").strip():
            issues.append(ValidationIssue("ERROR", "B1_Corpus_Documents", row, "publishing_org_id", "Included document row has no publishing_org_id."))
        if r.get("publishing_org_id", "").strip() and r.get("publishing_org_id", "").strip() not in org_ids:
            issues.append(ValidationIssue("ERROR", "B1_Corpus_Documents", row, "publishing_org_id", f"publishing_org_id {r.get('publishing_org_id')} does not exist in A1."))
        if r.get("file_name", "").strip() and not r.get("apa_reference", "").strip():
            issues.append(ValidationIssue("WARNING", "B1_Corpus_Documents", row, "apa_reference", "Document has no APA reference."))

    seen_variants = set()
    active_rows = 0
    for r in dictionary:
        row = r.get("__excel_row", "")
        active = r.get("active", "").strip().lower() in {"yes", "true", "1", "y"}
        if active:
            active_rows += 1
            for field in ["term_id", "canonical_term", "search_variant", "category"]:
                if not r.get(field, "").strip():
                    issues.append(ValidationIssue("ERROR", "Dictionary", row, field, "Active dictionary row is missing a required value."))
            key = (r.get("term_id", "").strip().lower(), r.get("search_variant", "").strip().lower())
            if key in seen_variants:
                issues.append(ValidationIssue("WARNING", "Dictionary", row, "search_variant", "Duplicate active term_id + search_variant combination."))
            seen_variants.add(key)
    if active_rows == 0:
        issues.append(ValidationIssue("ERROR", "Dictionary", "all", "active", "No active dictionary rows found."))

    excluded = {r.get("file_name", "").strip().lower() for r in exclusions if "exclude" in r.get("rule", "").lower()}
    for r in b1:
        fn = r.get("file_name", "").strip().lower()
        if fn and fn in excluded:
            issues.append(ValidationIssue("ERROR", "B1_Corpus_Documents", r.get("__excel_row", ""), "file_name", f"File {fn} is listed in Exclusions_Duplicates."))

    if pdf_folder and pdf_folder.exists():
        pdf_names = {p.name.lower() for p in pdf_folder.rglob("*.pdf")}
        zip_present = any(p.suffix.lower() == ".zip" for p in pdf_folder.rglob("*"))
        for r in b1:
            fn = r.get("file_name", "").strip()
            if fn and fn.lower().endswith(".pdf") and fn.lower() not in pdf_names and not zip_present:
                issues.append(ValidationIssue("WARNING", "B1_Corpus_Documents", r.get("__excel_row", ""), "file_name", f"Expected PDF not found in selected folder: {fn}"))
    return issues


def write_issues_csv(issues: List[ValidationIssue], out_csv: Path) -> None:
    out_csv.parent.mkdir(parents=True, exist_ok=True)
    with out_csv.open("w", newline="", encoding="utf-8") as f:
        writer = csv.DictWriter(f, fieldnames=["timestamp", "severity", "sheet", "row", "field", "message"])
        writer.writeheader()
        now = datetime.now().isoformat(timespec="seconds")
        for issue in issues:
            d = asdict(issue)
            d["timestamp"] = now
            writer.writerow(d)


def write_rows_csv(rows: List[Dict[str, str]], out_csv: Path) -> None:
    out_csv.parent.mkdir(parents=True, exist_ok=True)
    if not rows:
        out_csv.write_text("", encoding="utf-8")
        return
    fieldnames = [c for c in rows[0].keys() if c != "__excel_row"]
    with out_csv.open("w", newline="", encoding="utf-8") as f:
        writer = csv.DictWriter(f, fieldnames=fieldnames, extrasaction="ignore")
        writer.writeheader()
        writer.writerows(rows)


def export_runtime_inputs(data: Dict[str, List[Dict[str, str]]], out_dir: Path) -> Dict[str, Path]:
    out_dir.mkdir(parents=True, exist_ok=True)
    outputs: Dict[str, Path] = {}
    mapping = {
        "A1_Organizations": "A1_Organizations.csv",
        "B1_Corpus_Documents": "B1_Corpus_Documents.csv",
        "Dictionary": "keyword_dictionary_v1_3.csv",
        "Run_Settings": "Run_Settings.csv",
        "Exclusions_Duplicates": "Exclusions_Duplicates.csv",
        "New_Documents": "New_Documents.csv",
    }
    for sheet, filename in mapping.items():
        rows = data.get(sheet, [])
        path = out_dir / filename
        write_rows_csv(rows, path)
        outputs[sheet] = path
    return outputs


def _replace_sheet_with_rows(wb: Workbook, sheet_name: str, rows: List[Dict[str, str]]) -> None:
    if sheet_name in wb.sheetnames:
        del wb[sheet_name]
    ws = wb.create_sheet(sheet_name)
    if not rows:
        ws.append(["no_rows"])
        return
    headers = [c for c in rows[0].keys() if c != "__excel_row"]
    ws.append(headers)
    for r in rows:
        ws.append([r.get(h, "") for h in headers])


def create_runtime_baseline_workbook(input_data: Dict[str, List[Dict[str, str]]], baseline_v13: Path, out_xlsx: Path) -> Path:
    """Create a runtime v1.3 baseline workbook with user A1/B1 metadata injected."""
    out_xlsx.parent.mkdir(parents=True, exist_ok=True)
    shutil.copy2(baseline_v13, out_xlsx)
    wb = load_workbook(out_xlsx)
    b1_rows = input_data.get("B1_Corpus_Documents", [])
    a1_rows = input_data.get("A1_Organizations", [])
    _replace_sheet_with_rows(wb, "B1_Corpus_Metadata_Snapshot", b1_rows)
    _replace_sheet_with_rows(wb, "Vetted_Corpus_Metadata", b1_rows)
    _replace_sheet_with_rows(wb, "A1_Org_Links_Snapshot", a1_rows)
    wb.save(out_xlsx)
    return out_xlsx
