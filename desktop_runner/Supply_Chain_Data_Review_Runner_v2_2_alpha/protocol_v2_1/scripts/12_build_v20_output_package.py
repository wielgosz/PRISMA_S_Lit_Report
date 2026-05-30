#!/usr/bin/env python3
"""
12_build_v20_output_package.py
==============================

Stage 12 of the Supply Chain Data Review Protocol v2.0.

Purpose
-------
Assemble all validated outputs into a final downloadable package while
preserving backwards compatibility with the May 27 keyword-output contract.

This repaired version addresses export issues discovered during the parallel
v2.0 execution and folds in the clarified figure contract:

1. Excel workbook exports must be sanitized and validated so the resulting
   .xlsx file opens cleanly in Excel.
2. SVG outputs must preserve the established legacy naming contract used by
   earlier PRISMA-S / DCF runs.
3. Only one canonical set of SVG filenames should be emitted.
4. Figures 1-3 should plot document frequency (`reports_referencing`) rather
   than raw term-occurrence totals.
5. Figure 1 should contain jurisdictional / landscape terms only; AOI terms
   are excluded.

Design notes
------------
- Analytical CSV outputs remain canonical.
- Workbook generation is a rendering step over those CSVs.
- Keyword CSVs are emitted in generic names and, when possible, May 27
  compatibility alias names are also written.
"""
from __future__ import annotations

import argparse
import json
import math
import re
import shutil
import zipfile
from datetime import datetime, timezone
from pathlib import Path
from typing import Dict, List, Optional

import pandas as pd

try:
    from openpyxl import Workbook, load_workbook
    from openpyxl.cell.cell import ILLEGAL_CHARACTERS_RE
except Exception:  # pragma: no cover
    Workbook = None
    load_workbook = None
    ILLEGAL_CHARACTERS_RE = re.compile(r"[\x00-\x08]|[\x0B-\x0C]|[\x0E-\x1F]")

# Canonical SVG filename contract retained from prior protocol runs.
LEGACY_FIGURE_SPECS = [
    {
        "category": "Jurisdictional terms",
        "legacy_filename": "DCF_PRISMA_S_Figure_1_jurisdictional_terms.svg",
        "title": "Jurisdictional and landscape terms",
    },
    {
        "category": "Supply chain terms",
        "legacy_filename": "DCF_PRISMA_S_Figure_2_supply_chain_terms.svg",
        "title": "Supply chain terms",
    },
    {
        "category": "Farm level terms",
        "legacy_filename": "DCF_PRISMA_S_Figure_3_farm_level_terms.svg",
        "title": "Farm level terms",
    },
]

MAY27_KEYWORD_ALIASES = {
    "Document_Term_Matrix.csv": "test_current_document_term_matrix_after_batch11.csv",
    "Document_Term_Counts.csv": "test_document_term_counts_after_batch11.csv",
    "Term_Summary.csv": "test_term_summary_after_batch11.csv",
    "Zero_Reference_Terms.csv": "test_zero_reference_terms_after_batch11.csv",
}


def load_csv_if_exists(path: Path) -> Optional[pd.DataFrame]:
    return pd.read_csv(path) if path.exists() else None


def sanitize_excel_value(value):
    """Convert a cell value to an Excel-safe scalar."""
    if value is None:
        return ""
    try:
        if pd.isna(value):
            return ""
    except Exception:
        pass
    if isinstance(value, float) and math.isnan(value):
        return ""
    if isinstance(value, (dict, list, tuple, set)):
        try:
            value = json.dumps(list(value) if isinstance(value, set) else value, ensure_ascii=False)
        except Exception:
            value = str(value)
    elif not isinstance(value, (str, int, float, bool)):
        value = str(value)
    if isinstance(value, str):
        value = ILLEGAL_CHARACTERS_RE.sub("", value)
        value = value.replace("\x00", "")
    return value


def sanitize_dataframe(df: Optional[pd.DataFrame]) -> pd.DataFrame:
    if df is None:
        return pd.DataFrame()
    sdf = df.copy()
    sdf.columns = [str(c)[:255] for c in sdf.columns]
    for col in sdf.columns:
        sdf[col] = sdf[col].map(sanitize_excel_value)
    return sdf


def write_df_to_sheet(wb, sheet_name: str, df: pd.DataFrame) -> None:
    safe_name = sheet_name[:31]
    if safe_name in wb.sheetnames:
        ws = wb[safe_name]
        ws.delete_rows(1, ws.max_row)
    else:
        ws = wb.create_sheet(safe_name)
    sdf = sanitize_dataframe(df)
    ws.append(list(sdf.columns))
    for row in sdf.itertuples(index=False, name=None):
        ws.append(list(row))
    for col in ws.columns:
        letter = col[0].column_letter
        max_len = max([len(str(c.value)) if c.value is not None else 0 for c in col] + [8])
        ws.column_dimensions[letter].width = min(max_len + 2, 60)
    ws.freeze_panes = "A2"


def build_workbook(template: Path, out_xlsx: Path, sheets: Dict[str, pd.DataFrame]) -> None:
    if load_workbook is None:
        raise RuntimeError("openpyxl is required to build the Excel workbook.")
    if template.exists():
        wb = load_workbook(template)
    else:
        wb = Workbook()
        if "Sheet" in wb.sheetnames:
            wb["Sheet"].title = "README"
    for sheet_name, df in sheets.items():
        write_df_to_sheet(wb, sheet_name, df)
    wb.save(out_xlsx)
    # Validation pass: ensure the written workbook can be re-opened.
    test_wb = load_workbook(out_xlsx, read_only=True)
    _ = test_wb.sheetnames
    test_wb.close()


def make_legacy_svg_bar_chart(
    term_summary: pd.DataFrame,
    out_path: Path,
    category: str,
    title: str,
    shared_xlim: float,
    axis_label: str = "Number of reports referencing term",
    measure_column: str = "reports_referencing",
) -> None:
    """Render a legacy-compatible SVG bar chart.

    Figure contract for the updated protocol:
    - Only one canonical SVG filename set is emitted.
    - Figure 1 contains only jurisdictional / landscape terms. AOI terms are
      excluded by selecting only the "Jurisdictional terms" category.
    - Figures use document frequency (`reports_referencing`) rather than raw
      total term occurrences.
    - A shared x-axis scale is maintained across Figures 1-3.
    - Highest values appear at the top of the chart.
    """
    try:
        import matplotlib
        matplotlib.use("Agg")
        import matplotlib.pyplot as plt
    except Exception:
        out_path.write_text(f"<!-- matplotlib unavailable; could not generate {title} -->", encoding="utf-8")
        return

    df = sanitize_dataframe(term_summary)
    if df.empty or "term" not in df.columns or "category" not in df.columns:
        out_path.write_text(f"<!-- no term summary available for {title} -->", encoding="utf-8")
        return

    if measure_column in df.columns:
        value_col = measure_column
    elif "reports_referencing" in df.columns:
        value_col = "reports_referencing"
    elif "total_occurrences" in df.columns:
        value_col = "total_occurrences"
    else:
        value_col = "count"

    sub = df[df["category"].astype(str).str.lower() == category.lower()].copy()
    if sub.empty:
        out_path.write_text(f"<!-- no rows found for {title} -->", encoding="utf-8")
        return

    sub[value_col] = pd.to_numeric(sub[value_col], errors="coerce").fillna(0)
    sub = sub.sort_values([value_col, "term"], ascending=[False, True])
    data = sub.iloc[::-1].copy()  # reverse for barh so highest values display at top

    height = max(3.2, 0.32 * len(data) + 1.4)
    fig, ax = plt.subplots(figsize=(7.2, height))
    ax.barh(data["term"].astype(str), data[value_col], color="#F0B310", height=0.55)
    ax.set_xlim(0, shared_xlim if shared_xlim > 0 else 1)
    ax.set_xlabel(axis_label)
    ax.set_title(title)
    ax.spines["top"].set_visible(False)
    ax.spines["right"].set_visible(False)
    for i, v in enumerate(data[value_col].tolist()):
        ax.text(float(v) + max(shared_xlim, 1) * 0.01, i, str(int(v)), va="center", fontsize=8)
    plt.tight_layout()
    fig.savefig(out_path, format="svg")
    plt.close(fig)


def zip_dir(src: Path, zip_path: Path) -> None:
    if zip_path.exists():
        zip_path.unlink()
    with zipfile.ZipFile(zip_path, "w", compression=zipfile.ZIP_DEFLATED) as z:
        for p in sorted(src.rglob("*")):
            if p.is_file():
                z.write(p, p.relative_to(src))


def copy_csvs(src_dir: Path, dst_dir: Path, alias_map: Dict[str, str] | None = None) -> List[str]:
    dst_dir.mkdir(parents=True, exist_ok=True)
    written: List[str] = []
    alias_map = alias_map or {}
    for csv_path in sorted(src_dir.glob("*.csv")):
        target = dst_dir / csv_path.name
        shutil.copy2(csv_path, target)
        written.append(target.name)
        if csv_path.name in alias_map:
            alias_target = dst_dir / alias_map[csv_path.name]
            shutil.copy2(csv_path, alias_target)
            written.append(alias_target.name)
    return written


def maybe_sheet(df: Optional[pd.DataFrame]) -> pd.DataFrame:
    return sanitize_dataframe(df if df is not None else pd.DataFrame())


def main() -> int:
    ap = argparse.ArgumentParser(description="12 - Build final v2.0 workbook, SVGs, CSV bundle and manifest.")
    ap.add_argument("--template", required=True)
    ap.add_argument("--corpus-manifest", required=True)
    ap.add_argument("--keyword-dir", required=True)
    ap.add_argument("--dataset-dir", required=True, help="Root containing dataset stage outputs and C1 normalization outputs.")
    ap.add_argument("--qa-dir", required=True)
    ap.add_argument("--out-dir", required=True)
    args = ap.parse_args()

    out = Path(args.out_dir)
    out.mkdir(parents=True, exist_ok=True)
    keyword = Path(args.keyword_dir)
    dataset = Path(args.dataset_dir)
    qa = Path(args.qa_dir)

    corpus = load_csv_if_exists(Path(args.corpus_manifest))
    corpus = corpus if corpus is not None else pd.DataFrame()
    term_summary = load_csv_if_exists(keyword / "Term_Summary.csv")
    term_summary = term_summary if term_summary is not None else pd.DataFrame()
    zero = load_csv_if_exists(keyword / "Zero_Reference_Terms.csv")
    zero = zero if zero is not None else pd.DataFrame()
    counts = load_csv_if_exists(keyword / "Document_Term_Counts.csv")
    counts = counts if counts is not None else pd.DataFrame()
    matrix = load_csv_if_exists(keyword / "Document_Term_Matrix.csv")
    matrix = matrix if matrix is not None else pd.DataFrame()
    d1 = load_csv_if_exists(keyword / "D1_Key_Terms.csv")
    d1 = d1 if d1 is not None else term_summary

    c1_registry = load_csv_if_exists(dataset / "c1_normalized" / "dataset_canonical_registry_v2_0.csv")
    c1_registry = c1_registry if c1_registry is not None else pd.DataFrame()
    e1 = load_csv_if_exists(dataset / "c1_normalized" / "e1_subset_from_c1_v2_0.csv")
    e1 = e1 if e1 is not None else pd.DataFrame()
    stageA = load_csv_if_exists(dataset / "stageA" / "Dataset_Mentions_Raw.csv")
    stageA = stageA if stageA is not None else pd.DataFrame()
    stageB = load_csv_if_exists(dataset / "stageB" / "Dataset_Mentions_StageB_Mapped.csv")
    stageB = stageB if stageB is not None else pd.DataFrame()
    stageC = load_csv_if_exists(dataset / "stageC" / "Dataset_by_Document.csv")
    stageC = stageC if stageC is not None else pd.DataFrame()
    stageD = load_csv_if_exists(dataset / "stageD" / "Dataset_Summary_Ranking.csv")
    stageD = stageD if stageD is not None else pd.DataFrame()

    a1 = pd.DataFrame()
    b1 = pd.DataFrame()
    if not corpus.empty:
        b1_cols = [
            c for c in [
                "doc_id", "title", "year", "authors_or_orgs", "publishing_org_id",
                "source_url", "apa_reference", "file_name", "status", "status_reason"
            ] if c in corpus.columns
        ]
        b1 = corpus[b1_cols].copy() if b1_cols else pd.DataFrame()
        if "publishing_org_id" in corpus.columns:
            a1_cols = [c for c in ["publishing_org_id", "authors_or_orgs"] if c in corpus.columns]
            a1 = corpus[a1_cols].drop_duplicates().rename(
                columns={"publishing_org_id": "org_id", "authors_or_orgs": "organization_name"}
            )

    workbook_path = out / "Supply_Chain_Data_Review_Protocol_v2_0_outputs.xlsx"
    sheets = {
        "Table A1": maybe_sheet(a1),
        "Table B1": maybe_sheet(b1),
        "Table C1": maybe_sheet(c1_registry),
        "Table D1": maybe_sheet(d1),
        "Table E1": maybe_sheet(e1),
        "Corpus Manifest": maybe_sheet(corpus),
        "Zero Reference Terms": maybe_sheet(zero),
        "Term Summary": maybe_sheet(term_summary),
        "Document Term Counts": maybe_sheet(counts),
        "Document Term Matrix": maybe_sheet(matrix),
        "Dataset Mentions Raw": maybe_sheet(stageA),
        "Dataset Mentions Mapped": maybe_sheet(stageB),
        "Dataset by Document": maybe_sheet(stageC),
        "Dataset Summary": maybe_sheet(stageD),
    }
    build_workbook(Path(args.template), workbook_path, sheets)

    keyword_csv_dir = out / "keyword_csv"
    dataset_csv_dir = out / "dataset_csv"
    qa_csv_dir = out / "qa_csv"
    written_keyword = copy_csvs(keyword, keyword_csv_dir, MAY27_KEYWORD_ALIASES)
    written_dataset: List[str] = []
    for stage_dir in [dataset / "c1_normalized", dataset / "stageA", dataset / "stageB", dataset / "stageC", dataset / "stageD"]:
        if stage_dir.exists():
            written_dataset.extend(copy_csvs(stage_dir, dataset_csv_dir))
    written_qa = copy_csvs(qa, qa_csv_dir)

    fig_dir = out / "figures_svg"
    fig_dir.mkdir(exist_ok=True)
    shared_xlim = 0.0
    if not term_summary.empty:
        shared_series = None
        if "reports_referencing" in term_summary.columns:
            shared_series = pd.to_numeric(term_summary["reports_referencing"], errors="coerce").fillna(0)
        elif "total_occurrences" in term_summary.columns:
            shared_series = pd.to_numeric(term_summary["total_occurrences"], errors="coerce").fillna(0)
        if shared_series is not None and len(shared_series):
            shared_xlim = float(shared_series.max()) * 1.08

    for spec in LEGACY_FIGURE_SPECS:
        legacy_path = fig_dir / spec["legacy_filename"]
        make_legacy_svg_bar_chart(
            term_summary,
            legacy_path,
            spec["category"],
            spec["title"],
            shared_xlim,
            axis_label="Number of reports referencing term",
            measure_column="reports_referencing",
        )

    manifest = {
        "protocol_name": "Supply Chain Data Review Protocol",
        "protocol_version": "2.1",
        "package_export_status": "repaired_excel_and_svg_export",
        "built_at_utc": datetime.now(timezone.utc).isoformat(),
        "workbook": workbook_path.name,
        "keyword_rule": "exact_case_insensitive_regex_with_alphanumeric_boundaries; variants rolled up to canonical terms",
        "corpus_rows": int(len(corpus)),
        "term_summary_rows": int(len(term_summary)),
        "dataset_stageA_rows": int(len(stageA)),
        "dataset_stageC_rows": int(len(stageC)),
        "qa_dir": str(qa),
        "keyword_csv_files": written_keyword,
        "dataset_csv_files": written_dataset,
        "qa_csv_files": written_qa,
        "figure_svg_files": [p.name for p in sorted(fig_dir.glob("*.svg"))],
        "figure_measure": "reports_referencing",
        "figure_axis_label": "Number of reports referencing term",
        "figure_1_scope": "Jurisdictional and landscape terms only; AOI terms excluded.",
        "figure_filename_contract": "Only canonical legacy filenames are emitted.",
    }
    (out / "v20_run_manifest.json").write_text(json.dumps(manifest, indent=2), encoding="utf-8")
    (out / "README_EXPORT_REPAIRS.txt").write_text(
        "This export path sanitizes workbook cell content, validates the saved XLSX, emits only the canonical legacy DCF_PRISMA_S_Figure_* SVG filenames, and renders figures from reports_referencing counts. Figure 1 excludes AOI terms.\n",
        encoding="utf-8",
    )

    bundle = out.with_suffix(".zip")
    zip_dir(out, bundle)
    print(f"12 complete: workbook={workbook_path}; bundle={bundle}")
    return 0


if __name__ == "__main__":
    raise SystemExit(main())

# Version 2.1 note: figure outputs use reports_referencing and only canonical SVG names.
