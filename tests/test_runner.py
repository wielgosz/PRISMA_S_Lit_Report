"""End-to-end orchestrator behaviour (prisma_s.runner)."""

import json
import tempfile
import zipfile
from pathlib import Path

import openpyxl
import pandas as pd
import pytest

from prisma_s.runner import OUTPUT_COLUMNS, run_analysis


def _kw(tmp_path, *pairs):
    p = tmp_path / "kw.csv"
    p.write_text("group,term\n" + "\n".join(f"{g},{t}" for g, t in pairs), encoding="utf-8")
    return p


def test_empty_corpus_raises(tmp_path):
    (tmp_path / "docs").mkdir()
    with pytest.raises(ValueError, match="No .pdf or .docx"):
        run_analysis("b", tmp_path / "out.xlsx", input_path=tmp_path / "docs",
                     keyword_csv=_kw(tmp_path, ("C", "Coffee")), emit_citation=False)


def test_zero_term_dictionary_raises(tmp_path, make_docx):
    make_docx("some text", name="a.docx")
    kw = tmp_path / "empty.csv"
    kw.write_text("group,term\nC,\n", encoding="utf-8")
    with pytest.raises(ValueError):
        run_analysis("b", tmp_path / "out.xlsx", input_path=tmp_path,
                     keyword_csv=kw, emit_citation=False)


def test_three_sheets_and_columns(tmp_path, make_docx):
    make_docx("Coffee and Cocoa and Coffee", name="a.docx")
    out = tmp_path / "out" / "r.xlsx"
    df = run_analysis("b", out, input_path=tmp_path,
                      keyword_csv=_kw(tmp_path, ("C", "Coffee"), ("C", "Cocoa")),
                      emit_citation=False)
    assert list(df.columns) == OUTPUT_COLUMNS
    wb = openpyxl.load_workbook(out)
    assert wb.sheetnames == ["Long_AllTerms", "PRISMA-S_Compliance", "Run_Metadata"]
    counts = {r["Term"]: r["Count"] for _, r in df.iterrows()}
    assert counts["Coffee"] == 2 and counts["Cocoa"] == 1
    meta = json.loads((out.parent / "run_metadata.json").read_text())
    assert meta["documents_processed"] == 1
    assert (out.parent / "HOW_TO_CITE.txt").exists()


def test_year_is_single_dtype_with_mixed_docs(tmp_path, make_pdf):
    make_pdf(["Study from 2011 about Coffee"], name="dated.pdf")
    make_pdf(["Coffee with no year at all"], name="undated.pdf")
    out = tmp_path / "r.xlsx"
    df = run_analysis("b", out, input_path=tmp_path,
                      keyword_csv=_kw(tmp_path, ("C", "Coffee")), emit_citation=False)
    assert str(df["Year"].dtype) == "Int64"


def test_zip_temp_dir_is_cleaned(tmp_path, make_docx):
    d = make_docx("Coffee here", name="a.docx")
    archive = tmp_path / "corpus.zip"
    with zipfile.ZipFile(archive, "w") as z:
        z.write(d, "a.docx")
        z.writestr("__MACOSX/._a.docx", b"junk")
    before = set(Path(tempfile.gettempdir()).glob("prisma_s_zip_*"))
    df = run_analysis("b", tmp_path / "out.xlsx", input_path=archive,
                      keyword_csv=_kw(tmp_path, ("C", "Coffee")), emit_citation=False)
    after = set(Path(tempfile.gettempdir()).glob("prisma_s_zip_*"))
    assert before == after                      # no leak
    assert df["Document Name"].nunique() == 1   # __MACOSX entry ignored


def test_corrupt_pdf_counted_as_skipped(tmp_path, make_docx):
    make_docx("Coffee", name="good.docx")
    (tmp_path / "bad.pdf").write_bytes(b"%PDF-1.4 not really a pdf")
    out = tmp_path / "r.xlsx"
    run_analysis("b", out, input_path=tmp_path,
                 keyword_csv=_kw(tmp_path, ("C", "Coffee")), emit_citation=False)
    meta = json.loads((out.parent / "run_metadata.json").read_text())
    assert meta["documents_discovered"] == 2
    assert meta["documents_processed"] == 1
    assert meta["documents_skipped"] == 1


def test_run_metadata_has_needs_ocr_column_and_backend_counts(tmp_path, make_docx):
    make_docx("Coffee and Cocoa " * 40, name="a.docx")
    out = tmp_path / "r.xlsx"
    run_analysis("b", out, input_path=tmp_path,
                 keyword_csv=_kw(tmp_path, ("C", "Coffee")), emit_citation=False)
    meta = pd.read_excel(out, sheet_name="Run_Metadata")
    assert "Needs OCR" in meta.columns and "Escalated" not in meta.columns
    assert meta.loc[0, "Needs OCR"] in (False, 0)
    summary = json.loads((out.parent / "run_metadata.json").read_text())
    assert summary["backend_counts"].get("python-docx") == 1
    assert summary["documents_needing_ocr"] == []
    assert "ocr_enabled" not in summary and "escalated_documents" not in summary


def test_textless_pdf_surfaces_as_needs_ocr(tmp_path, monkeypatch, make_pdf):
    import prisma_s.extract as extract

    make_pdf(["one two three"], name="scan.pdf")

    def fake_text(path):
        return extract.ExtractResult(
            full_text="", first_text="", metadata={}, pages=12,
            backend="pypdf", thin=True, chain="pypdf 0w",
        )

    monkeypatch.setattr("prisma_s.runner.extract_text", fake_text)
    out = tmp_path / "r.xlsx"
    run_analysis("b", out, input_path=tmp_path,
                 keyword_csv=_kw(tmp_path, ("C", "Coffee")), emit_citation=False)
    meta = pd.read_excel(out, sheet_name="Run_Metadata")
    row = meta[meta["Document Name"] == "scan.pdf"].iloc[0]
    assert bool(row["Needs OCR"]) is True
    assert "OCR" in row["Status"]
    summary = json.loads((out.parent / "run_metadata.json").read_text())
    assert summary["documents_needing_ocr"] == ["scan.pdf"]


def test_repeated_runs_are_byte_stable(tmp_path, make_pdf):
    make_pdf(["Coffee and Cocoa on page one", "more Coffee on page two"], name="a.pdf")
    kw = _kw(tmp_path, ("C", "Coffee"), ("C", "Cocoa"))
    a = run_analysis("b", tmp_path / "a.xlsx", input_path=tmp_path, keyword_csv=kw,
                     emit_citation=False).drop(columns=["Run UTC"])
    b = run_analysis("b", tmp_path / "b.xlsx", input_path=tmp_path, keyword_csv=kw,
                     emit_citation=False).drop(columns=["Run UTC"])
    pd.testing.assert_frame_equal(a, b)


# ---------------------------------------------------------------------------
# v1.5 registry mode
# ---------------------------------------------------------------------------

def _reg(tmp_path):
    p = tmp_path / "reg.csv"
    p.write_text(
        "category,term_id,canonical_term,search_variant,active,include_in_visuals\n"
        "Jurisdictional terms,T1,area,area,yes,yes\n"
        "Jurisdictional terms,T1,area,sourcing area,yes,yes\n"
        "Supply chain terms,T2,supplier,supplier,yes,yes\n"
        "Supply chain terms,T3,mill,mill,yes,no\n"          # excluded from D1
        "Farm level terms,T4,rare,rare-term,yes,yes\n",     # zero references
        encoding="utf-8",
    )
    return p


def test_registry_mode_sheets_and_rollup(tmp_path, make_docx):
    make_docx("The area and the sourcing area. A supplier and a mill.", name="a.docx")
    out = tmp_path / "out" / "r.xlsx"
    df = run_analysis("v13", out, input_path=tmp_path, keyword_csv=_reg(tmp_path),
                      emit_citation=False, figures=False)

    assert list(df.columns)[:7] == [
        "Batch", "Document Name", "Title", "Year", "Category", "Term ID", "Canonical Term",
    ]
    assert (df["Protocol Version"] == "1.3").all()
    # roll-up = sum of variant counts: "area" x2 (incl. inside "sourcing area")
    # + "sourcing area" x1 = 3.  Overlapping variants summing is the paper's rule.
    area = df[df["Canonical Term"] == "area"].iloc[0]
    assert area["Count"] == 3 and area["Referenced"] == 1

    wb = openpyxl.load_workbook(out)
    assert wb.sheetnames == [
        "Long_AllTerms", "Term_Summary", "D1_Key_Terms", "Zero_Reference_Terms",
        "PRISMA-S_Compliance", "Run_Metadata",
    ]
    d1 = pd.read_excel(out, sheet_name="D1_Key_Terms")
    assert "mill" not in set(d1["term"])          # include_in_visuals == no
    assert "rare" not in set(d1["term"])          # zero references
    assert list(d1[d1["category"] == "Jurisdictional terms"]["term"]) == ["area"]
    zero = pd.read_excel(out, sheet_name="Zero_Reference_Terms")
    assert "rare" in set(zero["term"])

    meta = json.loads((out.parent / "run_metadata.json").read_text())
    assert meta["dictionary_mode"] == "registry"
    assert meta["n_canonical_terms"] == 4


def test_registry_runs_are_byte_stable(tmp_path, make_pdf):
    make_pdf(["area and sourcing area", "supplier and area again"], name="a.pdf")
    kw = _reg(tmp_path)
    a = run_analysis("b", tmp_path / "a.xlsx", input_path=tmp_path, keyword_csv=kw,
                     emit_citation=False, figures=False).drop(columns=["Run UTC"])
    b = run_analysis("b", tmp_path / "b.xlsx", input_path=tmp_path, keyword_csv=kw,
                     emit_citation=False, figures=False).drop(columns=["Run UTC"])
    pd.testing.assert_frame_equal(a, b)
